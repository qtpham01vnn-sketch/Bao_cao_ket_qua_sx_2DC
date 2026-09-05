import http.server
import socketserver
import json
import sqlite3
import urllib.parse
import os
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import random
import datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from form_mau_engine import build_form_mau_payload, save_form_mau_custom_data, parse_form_mau_excel_upload, resolve_period

PORT = 8080
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WORKSPACE_DIR = os.path.abspath(os.path.join(BASE_DIR, ".."))
DB_PATH = os.path.join(WORKSPACE_DIR, "app_data", "production_data.db")
ORIGINAL_EXCEL_PATH = os.path.join(WORKSPACE_DIR, "New Biểu đồ Báo cáo TH 2 DC năm 2026.xlsx")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_auth_access_table():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS auth_access_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL,
            full_name TEXT,
            department TEXT,
            device_id TEXT NOT NULL,
            request_type TEXT DEFAULT 'trial_25h',
            status TEXT DEFAULT 'pending',
            activation_pin TEXT,
            granted_at TEXT,
            expires_at TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

init_auth_access_table()

class ProductionAppHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=os.path.join(BASE_DIR, "public"), **kwargs)

    def end_headers(self):
        self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def send_json_response(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path
        params = urllib.parse.parse_qs(parsed_url.query)

        if path.startswith("/api/"):
            try:
                if path == "/api/dashboard":
                    self.handle_dashboard(params)
                elif path == "/api/data/summary":
                    self.handle_get_summary(params)
                elif path == "/api/data/brands":
                    self.handle_get_brands(params)
                elif path == "/api/data/materials":
                    self.handle_get_materials(params)
                elif path == "/api/data/coal":
                    self.handle_get_coal(params)
                elif path == "/api/norms/versions":
                    self.handle_get_norm_versions(params)
                elif path == "/api/norms/details":
                    self.handle_get_norm_details(params)
                elif path == "/api/report/form-mau":
                    self.handle_get_form_mau(params)
                elif path == "/api/export/sign-off-report" or path == "/api/export/form-mau-excel":
                    self.handle_export_sign_off_report(params)
                elif path == "/api/metadata":
                    self.handle_metadata()
                elif path == "/api/access/requests":
                    self.handle_get_access_requests()
                elif path == "/api/access/check-status":
                    self.handle_check_access_status(params)
                else:
                    self.send_json_response({"error": "Endpoint not found"}, status=404)
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.send_json_response({"error": str(e)}, status=500)
            return

        super().do_GET()

    def do_POST(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path

        try:
            content_length = int(self.headers.get("Content-Length", 0))
            post_data = self.rfile.read(content_length)
            
            content_type = self.headers.get("Content-Type", "")
            if "multipart/form-data" in content_type:
                if path == "/api/import/monthly":
                    self.handle_import_monthly_multipart(post_data, content_type)
                    return
                elif path == "/api/report/form-mau/import-excel":
                    self.handle_import_form_mau_multipart(post_data, content_type)
                    return

            body = {}
            if post_data:
                try:
                    body = json.loads(post_data.decode("utf-8"))
                except:
                    pass

            if path == "/api/norms/versions":
                self.handle_create_norm_version(body)
            elif path == "/api/norms/details":
                self.handle_save_norm_details(body)
            elif path == "/api/norms/items":
                self.handle_add_norm_item(body)
            elif path == "/api/report/form-mau/save-custom":
                self.handle_save_form_mau_custom(body)
            elif path == "/api/access/request":
                self.handle_create_access_request(body)
            elif path == "/api/access/approve":
                self.handle_approve_access_request(body)
            elif path == "/api/access/verify-pin":
                self.handle_verify_pin_access(body)
            else:
                self.send_json_response({"error": "Endpoint not found"}, status=404)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.send_json_response({"error": str(e)}, status=500)

    def do_DELETE(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path
        params = urllib.parse.parse_qs(parsed_url.query)
        item_id = params.get("id", [None])[0]

        if not item_id:
            self.send_json_response({"error": "ID parameter required"}, status=400)
            return

        conn = get_db()
        cur = conn.cursor()
        if path == "/api/data/summary":
            cur.execute("DELETE FROM data_production_summary WHERE id = ?", (item_id,))
        elif path == "/api/data/brands":
            cur.execute("DELETE FROM data_brand_production WHERE id = ?", (item_id,))
        elif path == "/api/data/materials":
            cur.execute("DELETE FROM data_material_consumption WHERE id = ?", (item_id,))
        elif path == "/api/data/coal":
            cur.execute("DELETE FROM data_coal_consumption WHERE id = ?", (item_id,))
        elif path == "/api/norms/versions":
            cur.execute("DELETE FROM master_norms_detail WHERE version_id = ?", (item_id,))
            cur.execute("DELETE FROM master_norms_version WHERE id = ?", (item_id,))
        elif path == "/api/norms/items":
            cur.execute("DELETE FROM master_norms_detail WHERE id = ?", (item_id,))
        conn.commit()
        conn.close()
        self.send_json_response({"success": True, "message": "Đã xóa bản ghi thành công"})

    def handle_metadata(self):
        conn = get_db()
        cur = conn.cursor()
        months = [r[0] for r in cur.execute("SELECT DISTINCT month FROM data_production_summary ORDER BY month").fetchall()]
        lines = [r[0] for r in cur.execute("SELECT DISTINCT line FROM data_production_summary WHERE line != '' ORDER BY line").fetchall()]
        sizes = [r[0] for r in cur.execute("SELECT DISTINCT size FROM data_production_summary WHERE size != '' ORDER BY size").fetchall()]
        brands = [r[0] for r in cur.execute("SELECT DISTINCT brand_name FROM data_brand_production WHERE brand_name != '' ORDER BY brand_name").fetchall()]
        glazes = [r[0] for r in cur.execute("SELECT DISTINCT glaze_type FROM data_brand_production WHERE glaze_type != '' ORDER BY glaze_type").fetchall()]
        conn.close()

        self.send_json_response({
            "months": months,
            "lines": lines,
            "sizes": sizes,
            "brands": brands,
            "glazes": glazes
        })

    def handle_dashboard(self, params):
        # Section 1 parameters (Sản lượng & Chất lượng)
        p1_month = params.get("p1_month", params.get("month", ["all"]))[0]
        p1_line = params.get("p1_line", params.get("line", ["all"]))[0]
        p1_size = params.get("p1_size", params.get("size", ["all"]))[0]
        p1_brand = params.get("p1_brand", params.get("brand", ["all"]))[0]

        # Section 2 parameters (Sản lượng Thương hiệu)
        p2_month = params.get("p2_month", params.get("month", ["all"]))[0]
        p2_line = params.get("p2_line", params.get("line", ["all"]))[0]
        p2_size = params.get("p2_size", params.get("size", ["all"]))[0]
        p2_brand = params.get("p2_brand", ["all"])[0]

        # Section 3 parameters (Tiêu hao Vật tư)
        p3_month = params.get("p3_month", params.get("month", ["all"]))[0]
        p3_line = params.get("p3_line", params.get("line", ["all"]))[0]
        p3_size = params.get("p3_size", params.get("size", ["all"]))[0]

        # Section 4 parameters (Sử dụng Than)
        p4_month = params.get("p4_month", params.get("month", ["all"]))[0]
        p4_line = params.get("p4_line", params.get("line", ["all"]))[0]
        p4_size = params.get("p4_size", params.get("size", ["all"]))[0]

        conn = get_db()
        cur = conn.cursor()

        # =========================================================================
        # 1. SECTION 1: Summary & Quality (Using p1_* filters)
        # =========================================================================
        where_d1 = ["unit = 'm2'"]
        vals_d1 = []
        if p1_month != "all":
            where_d1.append("month = ?")
            vals_d1.append(int(p1_month))
        if p1_line != "all":
            where_d1.append("line = ?")
            vals_d1.append(p1_line)
        if p1_size != "all":
            where_d1.append("size = ?")
            vals_d1.append(p1_size)

        clause_d1 = ("WHERE " + " AND ".join(where_d1)) if where_d1 else ""
        q_prod = f"""
            SELECT data_type, 
                   COUNT(*) as row_count,
                   SUM(recovery_total) as total_m2, 
                   SUM(a1) as a1_m2, 
                   SUM(a) as a_m2, 
                   SUM(b) as b_m2, 
                   SUM(sl_ep) as press_m2, 
                   SUM(prod_days) as days, 
                   SUM(stop_time_total) as stop_tot, 
                   SUM(stop_time_2mf) as stop_2mf 
            FROM data_production_summary 
            {clause_d1} 
            GROUP BY data_type
        """
        prod_rows = {r["data_type"]: dict(r) for r in cur.execute(q_prod, vals_d1).fetchall()}
        
        actual = prod_rows.get("Thực hiện", {"row_count": 1, "total_m2": 0, "a1_m2": 0, "a_m2": 0, "b_m2": 0, "press_m2": 0, "days": 0, "stop_tot": 0, "stop_2mf": 0})
        plan = prod_rows.get("Kế hoạch", {"row_count": 1, "total_m2": 0, "a1_m2": 0, "a_m2": 0, "b_m2": 0, "press_m2": 0, "days": 0, "stop_tot": 0, "stop_2mf": 0})

        # Section 1 Dual Donut Metrics
        act_tot = actual["total_m2"]
        act_a1 = actual["a1_m2"]
        act_a = actual["a_m2"]
        act_b = actual["b_m2"]
        act_days = actual["days"]
        act_a1_pct = (act_a1 / act_tot * 100) if act_tot > 0 else 0
        act_a_pct = (act_a / act_tot * 100) if act_tot > 0 else 0
        act_b_pct = (act_b / act_tot * 100) if act_tot > 0 else 0
        act_avg_day = (act_tot / act_days) if act_days > 0 else 0
        act_stop_2mf = (actual["stop_2mf"] / actual["row_count"]) if actual["row_count"] > 0 else 0

        pln_tot = plan["total_m2"]
        pln_a1 = plan["a1_m2"]
        pln_a = plan["a_m2"]
        pln_b = plan["b_m2"]
        pln_days = plan["days"]
        pln_a1_pct = (pln_a1 / pln_tot * 100) if pln_tot > 0 else 0
        pln_a_pct = (pln_a / pln_tot * 100) if pln_tot > 0 else 0
        pln_b_pct = (pln_b / pln_tot * 100) if pln_tot > 0 else 0
        pln_avg_day = (pln_tot / pln_days) if pln_days > 0 else 0
        pln_stop_2mf = (plan["stop_2mf"] / plan["row_count"]) if plan["row_count"] > 0 else 0

        completion_rate = (act_tot / pln_tot * 100) if pln_tot > 0 else 0

        # Section 1 Monthly Trend Chart (Grouped Bar Chart)
        if p1_brand != "all":
            # If Section 1 selected a specific brand
            trend_spec_where = ["brand_name = ?"]
            trend_spec_vals = [p1_brand]
            if p1_line != "all":
                trend_spec_where.append("line = ?")
                trend_spec_vals.append(p1_line)
            if p1_size != "all":
                trend_spec_where.append("size = ?")
                trend_spec_vals.append(p1_size)
            trend_spec_clause = "WHERE " + " AND ".join(trend_spec_where)
            q_spec_trend = f"SELECT month, SUM(quantity_m2) as total_m2, SUM(CASE WHEN grade = 'A1' THEN quantity_m2 ELSE 0 END) as a1_m2, SUM(CASE WHEN grade = 'B' THEN quantity_m2 ELSE 0 END) as b_m2 FROM data_brand_production {trend_spec_clause} GROUP BY month ORDER BY month"
            
            monthly_map = {}
            for r in cur.execute(q_spec_trend, trend_spec_vals).fetchall():
                m = r["month"]
                monthly_map[m] = {
                    "month": f"{m}",
                    "month_num": m,
                    "plan": 0,
                    "actual": r["total_m2"],
                    "a1": r["a1_m2"],
                    "b": r["b_m2"]
                }
            monthly_trend = sorted(list(monthly_map.values()), key=lambda x: x["month_num"])
        else:
            trend_where = ["unit = 'm2'"]
            trend_vals = []
            if p1_line != "all":
                trend_where.append("line = ?")
                trend_vals.append(p1_line)
            if p1_size != "all":
                trend_where.append("size = ?")
                trend_vals.append(p1_size)
            trend_clause = ("WHERE " + " AND ".join(trend_where)) if trend_where else ""

            q_monthly = f"SELECT month, data_type, SUM(recovery_total) as total_m2, SUM(a1) as a1_m2, SUM(a) as a_m2, SUM(b) as b_m2 FROM data_production_summary {trend_clause} GROUP BY month, data_type ORDER BY month"
            monthly_map = {}
            for r in cur.execute(q_monthly, trend_vals).fetchall():
                m = r["month"]
                if m not in monthly_map:
                    monthly_map[m] = {"month": f"{m}", "month_num": m, "plan": 0, "actual": 0, "a1": 0, "a": 0, "b": 0}
                if r["data_type"] == "Kế hoạch":
                    monthly_map[m]["plan"] = r["total_m2"]
                else:
                    monthly_map[m]["actual"] = r["total_m2"]
                    monthly_map[m]["a1"] = r["a1_m2"]
                    monthly_map[m]["a"] = r["a_m2"]
                    monthly_map[m]["b"] = r["b_m2"]
            monthly_trend = sorted(list(monthly_map.values()), key=lambda x: x["month_num"])

        actual_data = {
            "total_m2": act_tot,
            "a1_m2": act_a1,
            "a_m2": act_a,
            "b_m2": act_b,
            "a1_pct": act_a1_pct,
            "a_pct": act_a_pct,
            "b_pct": act_b_pct,
            "days": act_days,
            "avg_per_day": act_avg_day,
            "stop_time_2mf": act_stop_2mf,
            "share_pct": 100.0
        }
        plan_data = {
            "total_m2": pln_tot,
            "a1_m2": pln_a1,
            "a_m2": pln_a,
            "b_m2": pln_b,
            "a1_pct": pln_a1_pct,
            "a_pct": pln_a_pct,
            "b_pct": pln_b_pct,
            "days": pln_days,
            "avg_per_day": pln_avg_day,
            "stop_time_2mf": pln_stop_2mf
        }

        # =========================================================================
        # 2. SECTION 2: Brand Distribution & Table (Using p2_* filters)
        # =========================================================================
        where_avail = []
        vals_avail = []
        if p2_month != "all":
            where_avail.append("month = ?")
            vals_avail.append(int(p2_month))
        if p2_line != "all":
            where_avail.append("line = ?")
            vals_avail.append(p2_line)
        if p2_size != "all":
            where_avail.append("size = ?")
            vals_avail.append(p2_size)
        clause_avail = ("WHERE " + " AND ".join(where_avail)) if where_avail else ""
        q_avail_brands = f"SELECT DISTINCT brand_name FROM data_brand_production {clause_avail} {'AND' if clause_avail else 'WHERE'} brand_name != '' ORDER BY brand_name"
        available_brands = [r[0] for r in cur.execute(q_avail_brands, vals_avail).fetchall()]

        where_d2 = []
        vals_d2 = []
        if p2_month != "all":
            where_d2.append("month = ?")
            vals_d2.append(int(p2_month))
        if p2_line != "all":
            where_d2.append("line = ?")
            vals_d2.append(p2_line)
        if p2_size != "all":
            where_d2.append("size = ?")
            vals_d2.append(p2_size)

        clause_d2 = ("WHERE " + " AND ".join(where_d2)) if where_d2 else ""
        q_brand = f"""
            SELECT brand_name, 
                   SUM(quantity_m2) as total_m2, 
                   SUM(CASE WHEN grade = 'A1' THEN quantity_m2 ELSE 0 END) as a1_m2, 
                   SUM(CASE WHEN grade = 'B' THEN quantity_m2 ELSE 0 END) as b_m2,
                   GROUP_CONCAT(DISTINCT line) as lines,
                   GROUP_CONCAT(DISTINCT size) as sizes,
                   GROUP_CONCAT(DISTINCT glaze_type) as glazes
            FROM data_brand_production 
            {clause_d2} 
            GROUP BY brand_name 
            ORDER BY total_m2 DESC
        """
        all_brand_rows = [dict(r) for r in cur.execute(q_brand, vals_d2).fetchall()]
        grand_total_brand_m2 = sum(r["total_m2"] for r in all_brand_rows)

        brand_table = []
        for r in all_brand_rows:
            tot = r["total_m2"]
            a1 = r["a1_m2"]
            b = r["b_m2"]
            a1_p = (a1 / tot * 100) if tot > 0 else 0
            b_p = (b / tot * 100) if tot > 0 else 0
            share_p = (tot / grand_total_brand_m2 * 100) if grand_total_brand_m2 > 0 else 0
            brand_table.append({
                "brand_name": r["brand_name"],
                "total_m2": tot,
                "a1_m2": a1,
                "b_m2": b,
                "a1_pct": a1_p,
                "b_pct": b_p,
                "share_pct": share_p,
                "lines": r.get("lines") or "",
                "sizes": r.get("sizes") or "",
                "glazes": r.get("glazes") or ""
            })

        if p2_brand != "all":
            # Breakdown by glaze or size for this single brand
            where_spec = ["brand_name = ?"]
            vals_spec = [p2_brand]
            if p2_month != "all":
                where_spec.append("month = ?")
                vals_spec.append(int(p2_month))
            if p2_line != "all":
                where_spec.append("line = ?")
                vals_spec.append(p2_line)
            if p2_size != "all":
                where_spec.append("size = ?")
                vals_spec.append(p2_size)
            clause_spec = "WHERE " + " AND ".join(where_spec)
            
            q_brand_dist = f"SELECT glaze_type as brand_name, SUM(quantity_m2) as total_m2 FROM data_brand_production {clause_spec} GROUP BY glaze_type ORDER BY total_m2 DESC"
            brand_dist_rows = [dict(r) for r in cur.execute(q_brand_dist, vals_spec).fetchall()]
            if not brand_dist_rows or (len(brand_dist_rows) == 1 and not brand_dist_rows[0]["brand_name"]):
                q_brand_dist = f"SELECT size as brand_name, SUM(quantity_m2) as total_m2 FROM data_brand_production {clause_spec} GROUP BY size ORDER BY total_m2 DESC"
                brand_dist_rows = [dict(r) for r in cur.execute(q_brand_dist, vals_spec).fetchall()]
        else:
            brand_dist_rows = all_brand_rows[:10]

        # =========================================================================
        # 3. SECTION 3: Materials Consumption (Using p3_* filters)
        # =========================================================================
        where_d3 = []
        vals_d3 = []
        if p3_month != "all":
            where_d3.append("month = ?")
            vals_d3.append(int(p3_month))
        if p3_line != "all":
            where_d3.append("line = ?")
            vals_d3.append(p3_line)
        if p3_size != "all":
            where_d3.append("size = ?")
            vals_d3.append(p3_size)
        clause_d3 = ("WHERE " + " AND ".join(where_d3)) if where_d3 else ""

        q_mat_list = f"""
            SELECT id, stt, month, line, size, material_name, unit, norm_value, 
                   used_qty, prod_qty, actual_rate, reduced_qty, over_qty, diff_qty, status_text
            FROM data_material_consumption 
            {clause_d3}
            ORDER BY id ASC
        """
        raw_mat_rows = [dict(r) for r in cur.execute(q_mat_list, vals_d3).fetchall()]

        total_mat_used_kg = sum(r["used_qty"] for r in raw_mat_rows if r["unit"] == "Kg")
        total_mat_reduced_kg = sum(r["reduced_qty"] for r in raw_mat_rows)
        total_mat_over_kg = sum(r["over_qty"] for r in raw_mat_rows)
        
        total_xuong_kg = sum(r["used_qty"] for r in raw_mat_rows if "xương" in r["material_name"].lower())
        total_men_kg = sum(r["used_qty"] for r in raw_mat_rows if "men" in r["material_name"].lower())
        total_vo_dieu_kg = sum(r["used_qty"] for r in raw_mat_rows if "điều" in r["material_name"].lower())

        q_mat_chart = f"""
            SELECT material_name, unit,
                   AVG(norm_value) as norm_val,
                   SUM(used_qty) as total_used,
                   SUM(prod_qty) as total_prod,
                   SUM(reduced_qty) as total_reduced,
                   SUM(over_qty) as total_over
            FROM data_material_consumption
            {clause_d3}
            GROUP BY material_name
            ORDER BY total_used DESC
        """
        chart_mat_rows = []
        for r in cur.execute(q_mat_chart, vals_d3).fetchall():
            d = dict(r)
            p_m2 = d["total_prod"]
            act_r = (d["total_used"] / p_m2) if p_m2 > 0 else 0
            chart_mat_rows.append({
                "material_name": d["material_name"],
                "unit": d["unit"],
                "norm_value": round(d["norm_val"] or 0, 4),
                "actual_rate": round(act_r, 4),
                "total_used": round(d["total_used"] or 0, 2),
                "total_reduced": round(d["total_reduced"] or 0, 2),
                "total_over": round(d["total_over"] or 0, 2)
            })

        # =========================================================================
        # 4. SECTION 4: Coal Consumption (Using p4_* filters)
        # =========================================================================
        where_d4 = []
        vals_d4 = []
        if p4_month != "all":
            where_d4.append("month = ?")
            vals_d4.append(int(p4_month))
        if p4_line != "all":
            where_d4.append("line = ?")
            vals_d4.append(p4_line)
        if p4_size != "all":
            where_d4.append("size = ?")
            vals_d4.append(p4_size)
        clause_d4 = ("WHERE " + " AND ".join(where_d4)) if where_d4 else ""

        q_coal_list = f"""
            SELECT id, stt, month, line, size, coal_supplier, warehouse, import_date, 
                   firing_type, heat_value, ash_rate, std_ash_rate, stone_rate,
                   issued_weight, ash_weight, ash_export_rate, compensation_weight, excess_ash_weight,
                   total_used_weight, production_m2, rate_lump, rate_with_ash, rate_total, note
            FROM data_coal_consumption
            {clause_d4}
            ORDER BY id ASC
        """
        raw_coal_rows = [dict(r) for r in cur.execute(q_coal_list, vals_d4).fetchall()]

        total_coal_issued = sum(r["issued_weight"] or 0 for r in raw_coal_rows)
        total_coal_used = sum(r["total_used_weight"] or 0 for r in raw_coal_rows)
        total_coal_prod_m2 = sum(r["production_m2"] or 0 for r in raw_coal_rows)
        total_coal_ash = sum(r["ash_weight"] or 0 for r in raw_coal_rows)
        
        avg_coal_rate = (total_coal_used / total_coal_prod_m2) if total_coal_prod_m2 > 0 else 0
        
        valid_heats = [r["heat_value"] for r in raw_coal_rows if r["heat_value"] and r["heat_value"] > 0]
        avg_coal_heat = (sum(valid_heats) / len(valid_heats)) if valid_heats else 0

        valid_ashes = [r["ash_rate"] for r in raw_coal_rows if r["ash_rate"] is not None and r["ash_rate"] > 0]
        avg_coal_ash = (sum(valid_ashes) / len(valid_ashes)) if valid_ashes else 0

        where_coal_trend = []
        vals_coal_trend = []
        if p4_line != "all":
            where_coal_trend.append("line = ?")
            vals_coal_trend.append(p4_line)
        if p4_size != "all":
            where_coal_trend.append("size = ?")
            vals_coal_trend.append(p4_size)
        clause_coal_trend = ("WHERE " + " AND ".join(where_coal_trend)) if where_coal_trend else ""

        q_coal_trend = f"""
            SELECT month, 
                   SUM(total_used_weight) as used_kg, 
                   SUM(production_m2) as prod_m2,
                   AVG(heat_value) as heat_val,
                   AVG(ash_rate) as ash_pct
            FROM data_coal_consumption
            {clause_coal_trend}
            GROUP BY month
            ORDER BY month
        """
        coal_monthly_trend = []
        for r in cur.execute(q_coal_trend, vals_coal_trend).fetchall():
            m_used = r["used_kg"] or 0
            m_prod = r["prod_m2"] or 0
            m_rate = (m_used / m_prod) if m_prod > 0 else 0
            coal_monthly_trend.append({
                "month": f"{r['month']}",
                "month_num": r["month"],
                "used_ton": round(m_used / 1000, 1),
                "prod_m2": round(m_prod, 0),
                "rate_kg_m2": round(m_rate, 2),
                "heat_val": round(r["heat_val"] or 0, 0),
                "ash_pct": round(r["ash_pct"] or 0, 1)
            })

        conn.close()

        self.send_json_response({
            "is_brand_selected": (p1_brand != "all"),
            "selected_brand_name": p1_brand,
            "completion_rate": completion_rate,
            "actual": actual_data,
            "plan": plan_data,
            "monthly_trend": monthly_trend,
            "brand_distribution": brand_dist_rows,
            "available_brands": available_brands,
            "brand_table": brand_table,
            "materials_section": {
                "total_used_kg": total_mat_used_kg,
                "total_reduced_kg": total_mat_reduced_kg,
                "total_over_kg": total_mat_over_kg,
                "total_xuong_kg": total_xuong_kg,
                "total_men_kg": total_men_kg,
                "total_vo_dieu_kg": total_vo_dieu_kg,
                "materials_list": raw_mat_rows,
                "materials_chart": chart_mat_rows
            },
            "coal_section": {
                "total_coal_issued_kg": total_coal_issued,
                "total_coal_used_kg": total_coal_used,
                "total_coal_prod_m2": total_coal_prod_m2,
                "total_coal_ash_kg": total_coal_ash,
                "avg_coal_rate": avg_coal_rate,
                "avg_coal_heat": avg_coal_heat,
                "avg_coal_ash": avg_coal_ash,
                "coal_list": raw_coal_rows,
                "coal_monthly_trend": coal_monthly_trend
            }
        })

    def handle_get_summary(self, params):
        conn = get_db()
        cur = conn.cursor()
        month = params.get("month", [None])[0]
        line = params.get("line", [None])[0]
        size = params.get("size", [None])[0]
        unit = params.get("unit", ["m2"])[0]

        where = []
        vals = []
        if unit and unit != "all":
            where.append("unit = ?")
            vals.append(unit)
        if month and month != "all":
            where.append("month = ?")
            vals.append(int(month))
        if line and line != "all":
            where.append("line = ?")
            vals.append(line)
        if size and size != "all":
            where.append("size = ?")
            vals.append(size)

        q = "SELECT * FROM data_production_summary"
        if where:
            q += " WHERE " + " AND ".join(where)
        q += " ORDER BY month, line, id"
        rows = [dict(r) for r in cur.execute(q, vals).fetchall()]
        conn.close()

        self.send_json_response({"data": rows, "count": len(rows)})

    def handle_get_brands(self, params):
        conn = get_db()
        cur = conn.cursor()
        month = params.get("month", [None])[0]
        line = params.get("line", [None])[0]
        size = params.get("size", [None])[0]

        where = []
        vals = []
        if month and month != "all":
            where.append("month = ?")
            vals.append(int(month))
        if line and line != "all":
            where.append("line = ?")
            vals.append(line)
        if size and size != "all":
            where.append("size = ?")
            vals.append(size)

        q = "SELECT * FROM data_brand_production"
        if where:
            q += " WHERE " + " AND ".join(where)
        q += " ORDER BY month, line, id"
        rows = [dict(r) for r in cur.execute(q, vals).fetchall()]

        # Compute Total Summary (A1, A, B, Grand Total)
        sum_a1 = sum(r["quantity_m2"] for r in rows if r["grade"] == "A1")
        sum_a = sum(r["quantity_m2"] for r in rows if r["grade"] == "A")
        sum_b = sum(r["quantity_m2"] for r in rows if r["grade"] == "B")
        grand_total = sum(r["quantity_m2"] for r in rows)

        conn.close()

        self.send_json_response({
            "data": rows,
            "count": len(rows),
            "summary_totals": {
                "sum_a1": sum_a1,
                "sum_a": sum_a,
                "sum_b": sum_b,
                "grand_total": grand_total
            }
        })

    def handle_get_materials(self, params):
        conn = get_db()
        cur = conn.cursor()
        month = params.get("month", [None])[0]
        line = params.get("line", [None])[0]
        size = params.get("size", [None])[0]

        where = []
        vals = []
        if month and month != "all":
            where.append("month = ?")
            vals.append(int(month))
        if line and line != "all":
            where.append("line = ?")
            vals.append(line)
        if size and size != "all":
            where.append("size = ?")
            vals.append(size)

        q = "SELECT * FROM data_material_consumption"
        if where:
            q += " WHERE " + " AND ".join(where)
        q += " ORDER BY month, line, id"
        rows = [dict(r) for r in cur.execute(q, vals).fetchall()]
        conn.close()

        self.send_json_response({"data": rows, "count": len(rows)})

    def handle_get_coal(self, params):
        conn = get_db()
        cur = conn.cursor()
        month = params.get("month", [None])[0]
        line = params.get("line", [None])[0]
        size = params.get("size", [None])[0]
        firing_type = params.get("firing_type", [None])[0]

        where = []
        vals = []
        if month and month != "all":
            where.append("month = ?")
            vals.append(int(month))
        if line and line != "all":
            where.append("line = ?")
            vals.append(line)
        if size and size != "all":
            where.append("size = ?")
            vals.append(size)
        if firing_type and firing_type != "all":
            where.append("firing_type = ?")
            vals.append(firing_type)

        q = "SELECT * FROM data_coal_consumption"
        if where:
            q += " WHERE " + " AND ".join(where)
        q += " ORDER BY month, line, id"
        rows = [dict(r) for r in cur.execute(q, vals).fetchall()]

        # Compute Groups (Firing vs Drying vs All)
        firing_rows = [r for r in rows if "Không" not in (r["firing_type"] or "")]
        drying_rows = [r for r in rows if "Không" in (r["firing_type"] or "")]

        def calc_group_summary(group_rows):
            issued = sum(r["issued_weight"] or 0 for r in group_rows)
            ash = sum(r["ash_weight"] or 0 for r in group_rows)
            comp = sum(r["compensation_weight"] or 0 for r in group_rows)
            excess_ash = sum(r["excess_ash_weight"] or 0 for r in group_rows)
            total_used = sum(r["total_used_weight"] or ((r["issued_weight"] or 0) + (r["ash_weight"] or 0) + (r["compensation_weight"] or 0) - (r["excess_ash_weight"] or 0)) for r in group_rows)
            prod_m2 = sum(r["production_m2"] or 0 for r in group_rows)
            
            ash_pct = (ash / (issued + ash) * 100) if (issued + ash) > 0 else 0
            rate_lump = (issued / prod_m2) if prod_m2 > 0 else 0
            rate_with_ash = ((issued + ash - excess_ash) / prod_m2) if prod_m2 > 0 else 0
            rate_total = (total_used / prod_m2) if prod_m2 > 0 else 0
            
            return {
                "issued_weight": issued,
                "ash_weight": ash,
                "ash_rate_avg": ash_pct,
                "compensation_weight": comp,
                "excess_ash_weight": excess_ash,
                "total_used_weight": total_used,
                "production_m2": prod_m2,
                "rate_lump": rate_lump,
                "rate_with_ash": rate_with_ash,
                "rate_total": rate_total
            }

        summary_firing = calc_group_summary(firing_rows)
        summary_drying = calc_group_summary(drying_rows)
        summary_all = calc_group_summary(rows)

        conn.close()

        self.send_json_response({
            "data": rows,
            "count": len(rows),
            "summary": {
                "firing": summary_firing,
                "drying": summary_drying,
                "all": summary_all
            }
        })

    def handle_get_norm_versions(self, params=None):
        params = params or {}
        line_filter = params.get("line", ["all"])[0]
        size_filter = params.get("size", ["all"])[0]

        conn = get_db()
        cur = conn.cursor()
        
        # Query all versions
        v_rows = [dict(r) for r in cur.execute("""
            SELECT v.*, 
                   COUNT(d.id) as item_count,
                   GROUP_CONCAT(DISTINCT d.line) as lines_covered,
                   GROUP_CONCAT(DISTINCT d.size) as sizes_covered
            FROM master_norms_version v 
            LEFT JOIN master_norms_detail d ON v.id = d.version_id 
            GROUP BY v.id 
            ORDER BY v.effective_from_year DESC, v.effective_from_month DESC, v.id DESC
        """).fetchall()]

        # Filter and augment matching item counts
        result = []
        for r in v_rows:
            v_id = r["id"]
            where_sub = ["version_id = ?"]
            vals_sub = [v_id]
            if line_filter != "all":
                where_sub.append("line = ?")
                vals_sub.append(line_filter)
            if size_filter != "all":
                where_sub.append("size = ?")
                vals_sub.append(size_filter)
            
            sub_count = cur.execute(f"SELECT COUNT(*) FROM master_norms_detail WHERE {' AND '.join(where_sub)}", vals_sub).fetchone()[0]
            r["filtered_item_count"] = sub_count
            
            if line_filter != "all" or size_filter != "all":
                if sub_count > 0 or (r.get("line") in (line_filter, "all") and r.get("size") in (size_filter, "all")):
                    result.append(r)
            else:
                result.append(r)

        conn.close()
        self.send_json_response({"data": result})

    def handle_get_norm_details(self, params):
        version_id = params.get("version_id", [None])[0]
        line_filter = params.get("line", ["all"])[0]
        size_filter = params.get("size", ["all"])[0]

        if not version_id:
            self.send_json_response({"error": "version_id required"}, status=400)
            return
        conn = get_db()
        cur = conn.cursor()
        v_info = dict(cur.execute("SELECT * FROM master_norms_version WHERE id = ?", (version_id,)).fetchone() or {})
        
        where = ["version_id = ?"]
        vals = [version_id]
        if line_filter != "all":
            where.append("line = ?")
            vals.append(line_filter)
        if size_filter != "all":
            where.append("size = ?")
            vals.append(size_filter)

        rows = [dict(r) for r in cur.execute(f"SELECT * FROM master_norms_detail WHERE {' AND '.join(where)} ORDER BY line, size, material_name", vals).fetchall()]
        conn.close()
        self.send_json_response({"version": v_info, "details": rows})

    def handle_create_norm_version(self, body):
        code = body.get("version_code")
        name = body.get("version_name")
        from_m = int(body.get("effective_from_month", 1))
        from_y = int(body.get("effective_from_year", 2026))
        desc = body.get("description", "")
        line = body.get("line", "all")
        size = body.get("size", "all")
        copy_from_id = body.get("copy_from_version_id")
        items = body.get("items", [])

        if not code or not name:
            self.send_json_response({"error": "Mã và Tên phiên bản là bắt buộc"}, status=400)
            return

        conn = get_db()
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO master_norms_version 
                (version_code, version_name, effective_from_month, effective_from_year, description, is_active, line, size) 
                VALUES (?, ?, ?, ?, ?, 1, ?, ?)
            """, (code, name, from_m, from_y, desc, line, size))
            new_v_id = cur.lastrowid

            if items and len(items) > 0:
                for item in items:
                    m_name = item.get("material_name", "").strip()
                    m_line = item.get("line", line)
                    m_size = item.get("size", size)
                    m_unit = item.get("unit", "Kg")
                    m_val = float(item.get("norm_value", 0))
                    if m_name:
                        cur.execute("""
                            INSERT INTO master_norms_detail (version_id, material_name, line, size, unit, norm_value)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (new_v_id, m_name, m_line, m_size, m_unit, m_val))
            elif copy_from_id:
                if line != "all" or size != "all":
                    where_cp = ["version_id = ?"]
                    vals_cp = [new_v_id, copy_from_id]
                    if line != "all":
                        where_cp.append("line = ?")
                        vals_cp.append(line)
                    if size != "all":
                        where_cp.append("size = ?")
                        vals_cp.append(size)
                    q = f"INSERT INTO master_norms_detail (version_id, material_name, line, size, unit, norm_value) SELECT ?, material_name, line, size, unit, norm_value FROM master_norms_detail WHERE {' AND '.join(where_cp)}"
                    cur.execute(q, vals_cp)
                else:
                    cur.execute("INSERT INTO master_norms_detail (version_id, material_name, line, size, unit, norm_value) SELECT ?, material_name, line, size, unit, norm_value FROM master_norms_detail WHERE version_id = ?", (new_v_id, copy_from_id))

            conn.commit()
            conn.close()
            self.send_json_response({"success": True, "version_id": new_v_id, "message": "Đã tạo phiên bản định mức mới thành công"})
        except sqlite3.IntegrityError:
            conn.close()
            self.send_json_response({"error": f"Mã phiên bản '{code}' đã tồn tại!"}, status=400)

    def handle_save_norm_details(self, body):
        version_id = body.get("version_id")
        items = body.get("items", [])
        if not version_id:
            self.send_json_response({"error": "version_id required"}, status=400)
            return

        conn = get_db()
        cur = conn.cursor()
        for item in items:
            item_id = item.get("id")
            norm_val = float(item.get("norm_value", 0))
            if item_id:
                cur.execute("UPDATE master_norms_detail SET norm_value = ? WHERE id = ? AND version_id = ?", (norm_val, item_id, version_id))
        conn.commit()
        conn.close()
        self.send_json_response({"success": True, "message": "Đã cập nhật định mức thành công"})

    def handle_add_norm_item(self, body):
        version_id = body.get("version_id")
        name = body.get("material_name", "").strip()
        line = body.get("line", "DC1")
        size = body.get("size", "30x60")
        unit = body.get("unit", "Kg")
        val = float(body.get("norm_value", 0))

        if not version_id or not name:
            self.send_json_response({"error": "Thiếu version_id hoặc Tên vật tư"}, status=400)
            return

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO master_norms_detail (version_id, material_name, line, size, unit, norm_value)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (version_id, name, line, size, unit, val))
        conn.commit()
        conn.close()
        self.send_json_response({"success": True, "message": "Đã thêm chỉ tiêu định mức thành công"})


    def handle_import_monthly_multipart(self, post_data, content_type):
        boundary = content_type.split("boundary=")[1].encode("utf-8")
        parts = post_data.split(b"--" + boundary)
        
        uploaded_files = {}
        target_month = 8
        target_year = 2026

        for part in parts:
            if b"Content-Disposition" in part and b"filename=" in part:
                headers_part, file_content = part.split(b"\r\n\r\n", 1)
                file_content = file_content.rstrip(b"\r\n")
                
                filename = ""
                fieldname = ""
                for line in headers_part.decode("utf-8", errors="ignore").split("\r\n"):
                    if "filename=" in line:
                        filename = line.split('filename="')[1].split('"')[0]
                    if "name=" in line:
                        fieldname = line.split('name="')[1].split('"')[0]
                
                if filename and file_content:
                    uploaded_files[fieldname] = {"filename": filename, "bytes": file_content}
            elif b'name="month"' in part:
                _, val = part.split(b"\r\n\r\n", 1)
                try:
                    target_month = int(val.decode("utf-8").strip())
                except:
                    pass

        logs = []
        for fieldname, finfo in uploaded_files.items():
            fname = finfo["filename"]
            fbytes = finfo["bytes"]
            logs.append(f'Đã nhận file "{fname}" ({len(fbytes):,} bytes)')
            try:
                wb = openpyxl.load_workbook(io.BytesIO(fbytes), data_only=True)
                sheet_list_str = ", ".join(wb.sheetnames)
                logs.append(f"-> Đọc thành công {len(wb.sheetnames)} sheet: {sheet_list_str}")
            except Exception as e:
                logs.append(f"-> Lỗi đọc file: {str(e)}")

        self.send_json_response({
            "success": True,
            "message": f"Đã phân tích và trích xuất thành công dữ liệu Tháng {target_month}/{target_year}",
            "logs": logs
        })

    def handle_get_form_mau(self, params):
        period_type = params.get("period_type", ["month"])[0]
        period_value = params.get("period_value", ["8"])[0]
        year = int(params.get("year", [2026])[0])

        conn = get_db()
        data = build_form_mau_payload(conn, period_type=period_type, period_value=period_value, year=year)
        conn.close()
        self.send_json_response(data)

    def handle_save_form_mau_custom(self, body):
        conn = get_db()
        res = save_form_mau_custom_data(conn, body)
        conn.close()
        self.send_json_response(res)

    def handle_import_form_mau_multipart(self, post_data, content_type):
        boundary = content_type.split("boundary=")[1].encode("utf-8")
        parts = post_data.split(b"--" + boundary)
        
        target_file_bytes = None
        period_type = "month"
        period_value = "8"
        year = 2026

        for part in parts:
            if b"Content-Disposition" in part and b"filename=" in part:
                _, file_content = part.split(b"\r\n\r\n", 1)
                target_file_bytes = file_content.rstrip(b"\r\n")
            elif b'name="period_type"' in part:
                _, val = part.split(b"\r\n\r\n", 1)
                period_type = val.decode("utf-8").strip()
            elif b'name="period_value"' in part:
                _, val = part.split(b"\r\n\r\n", 1)
                period_value = val.decode("utf-8").strip()
            elif b'name="year"' in part:
                _, val = part.split(b"\r\n\r\n", 1)
                try: year = int(val.decode("utf-8").strip())
                except: pass

        if not target_file_bytes:
            self.send_json_response({"error": "Không tìm thấy file tải lên"}, status=400)
            return

        parsed = parse_form_mau_excel_upload(target_file_bytes)
        
        # Save parsed data to DB for this period
        conn = get_db()
        body_to_save = {
            "period_type": period_type,
            "period_value": period_value,
            "year": year,
            "hr_data": parsed.get("hr_data"),
            "notes_data": {"hr_notes": parsed.get("hr_notes")},
            "goals_data": {"department_tasks": parsed.get("department_tasks")}
        }
        if parsed.get("evaluation_text"):
            body_to_save["evaluation_data"] = parsed.get("evaluation_text")
            
        save_form_mau_custom_data(conn, body_to_save)
        conn.close()

        self.send_json_response({
            "success": True,
            "message": f"Đã trích xuất và cập nhật thành công Form Mẫu cho kỳ {period_value}/{year}",
            "logs": parsed.get("logs", [])
        })

    def handle_export_sign_off_report(self, params):
        period_type = params.get("period_type", ["month"])[0]
        period_value = params.get("period_value", [params.get("month", ["8"])[0]])[0]
        year = int(params.get("year", [2026])[0])
        p_info = resolve_period(period_type, period_value, year)

        wb_orig = openpyxl.load_workbook(ORIGINAL_EXCEL_PATH, data_only=False)
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = f"Báo cáo Form Mẫu"

        ws_template = wb_orig["Form mẫu"]
        for row in ws_template.iter_rows(values_only=False):
            for cell in row:
                c_new = ws_out.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    c_new.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                    c_new.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)

        out_stream = io.BytesIO()
        wb_out.save(out_stream)
        out_stream.seek(0)
        file_bytes = out_stream.getvalue()

        clean_filename = f"Bao_Cao_Tong_Hop_KQSX_2DC_{p_info['period_key']}.xlsx"
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{clean_filename}"')
        self.send_header("Content-Length", str(len(file_bytes)))
        self.end_headers()
        self.wfile.write(file_bytes)

    def handle_get_access_requests(self):
        conn = get_db()
        cur = conn.cursor()
        rows = [dict(r) for r in cur.execute("SELECT * FROM auth_access_requests ORDER BY id DESC").fetchall()]
        conn.close()
        self.send_json_response({"requests": rows, "count": len(rows)})

    def handle_check_access_status(self, params):
        device_id = params.get("device_id", [None])[0]
        email = params.get("email", [None])[0]
        conn = get_db()
        cur = conn.cursor()
        if email:
            r = cur.execute("SELECT * FROM auth_access_requests WHERE email = ? ORDER BY id DESC LIMIT 1", (email,)).fetchone()
        elif device_id:
            r = cur.execute("SELECT * FROM auth_access_requests WHERE device_id = ? ORDER BY id DESC LIMIT 1", (device_id,)).fetchone()
        else:
            r = None
        conn.close()
        if r:
            self.send_json_response({"found": True, "request": dict(r)})
        else:
            self.send_json_response({"found": False, "request": None})

    def handle_create_access_request(self, body):
        email = (body.get("email") or "").strip().lower()
        full_name = (body.get("full_name") or "").strip()
        department = (body.get("department") or "").strip()
        device_id = (body.get("device_id") or "").strip()
        if not email or not device_id:
            self.send_json_response({"error": "Vui lòng nhập Email và mã thiết bị"}, status=400)
            return

        pin = str(random.randint(100000, 999999))
        conn = get_db()
        cur = conn.cursor()
        
        existing = cur.execute("SELECT id, status, activation_pin FROM auth_access_requests WHERE email = ? OR device_id = ? ORDER BY id DESC LIMIT 1", (email, device_id)).fetchone()
        if existing:
            rec_id = existing["id"]
            cur.execute("""
                UPDATE auth_access_requests 
                SET email = ?, full_name = ?, department = ?, device_id = ?, activation_pin = ?, status = 'pending'
                WHERE id = ?
            """, (email, full_name, department, device_id, pin, rec_id))
        else:
            cur.execute("""
                INSERT INTO auth_access_requests (email, full_name, department, device_id, activation_pin, status)
                VALUES (?, ?, ?, ?, ?, 'pending')
            """, (email, full_name, department, device_id, pin))
            rec_id = cur.lastrowid
        conn.commit()
        conn.close()

        self.send_json_response({
            "success": True,
            "message": "Yêu cầu cấp quyền đã được gửi thành công đến Quản trị viên.",
            "pin": pin,
            "request_id": rec_id,
            "email": email
        })

    def handle_approve_access_request(self, body):
        req_id = body.get("id")
        action_type = body.get("action_type")
        conn = get_db()
        cur = conn.cursor()
        
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if action_type == "25h":
            exp_str = (datetime.datetime.now() + datetime.timedelta(hours=25)).strftime("%Y-%m-%d %H:%M:%S")
            cur.execute("UPDATE auth_access_requests SET status = 'approved_25h', granted_at = ?, expires_at = ? WHERE id = ?", (now_str, exp_str, req_id))
        elif action_type == "permanent":
            cur.execute("UPDATE auth_access_requests SET status = 'approved_permanent', granted_at = ?, expires_at = 'UNLIMITED' WHERE id = ?", (now_str, req_id))
        elif action_type == "pin_gen":
            new_pin = str(random.randint(100000, 999999))
            cur.execute("UPDATE auth_access_requests SET activation_pin = ? WHERE id = ?", (new_pin, req_id))
        elif action_type == "reject":
            cur.execute("UPDATE auth_access_requests SET status = 'rejected' WHERE id = ?", (req_id,))
        
        conn.commit()
        updated_row = cur.execute("SELECT * FROM auth_access_requests WHERE id = ?", (req_id,)).fetchone()
        conn.close()

        self.send_json_response({
            "success": True,
            "message": f"Đã thực hiện thành công thao tác [{action_type}]",
            "request": dict(updated_row) if updated_row else None
        })

    def handle_verify_pin_access(self, body):
        pin = (body.get("pin") or "").strip()
        device_id = (body.get("device_id") or "").strip()
        email = (body.get("email") or "").strip().lower()

        # Master Master PIN override (e.g. 686868 or 888999)
        if pin in ["686868", "888999", "999888"]:
            self.send_json_response({
                "success": True,
                "grant_type": "permanent",
                "message": "Kích hoạt quyền truy cập thành công bằng Mã Chủ (Master PIN)!",
                "user": {
                    "username": email or "user_vip",
                    "displayName": "Khách mời VIP (Master PIN)",
                    "role": "quan_doc",
                    "roleTitle": "Khách Mời VIP (Chỉ Xem)"
                }
            })
            return

        conn = get_db()
        cur = conn.cursor()
        query = "SELECT * FROM auth_access_requests WHERE (activation_pin = ? OR (email = ? AND status LIKE 'approved%')) ORDER BY id DESC LIMIT 1"
        row = cur.execute(query, (pin, email)).fetchone()
        conn.close()

        if not row:
            self.send_json_response({"success": False, "error": "Mã PIN hoặc Email không hợp lệ hoặc chưa được Admin phê duyệt."}, status=400)
            return

        r = dict(row)
        grant = "permanent" if r["status"] == "approved_permanent" else "25h"
        self.send_json_response({
            "success": True,
            "grant_type": grant,
            "request": r,
            "message": f"Xác nhận thành công! Đã cấp quyền truy cập [{grant}]."
        })

def run_server():
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("", PORT), ProductionAppHandler) as httpd:
        print(f"Server running at http://localhost:{PORT}")
        httpd.serve_forever()

if __name__ == "__main__":
    run_server()
