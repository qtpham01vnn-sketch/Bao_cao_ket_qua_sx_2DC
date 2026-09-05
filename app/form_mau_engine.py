import sqlite3
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def init_form_mau_db(conn):
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS report_form_mau_custom (
            period_key TEXT PRIMARY KEY,
            period_type TEXT,
            period_value TEXT,
            year INTEGER DEFAULT 2026,
            hr_data TEXT,
            plan_data TEXT,
            goals_data TEXT,
            notes_data TEXT,
            evaluation_data TEXT,
            signatures_data TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()

def resolve_period(period_type="month", period_value="8", year=2026):
    if period_type == "month":
        m_int = int(period_value) if str(period_value).isdigit() else 8
        months = [m_int]
        period_title = f"THÁNG {m_int:02d} NĂM {year}"
        period_key = f"M{m_int:02d}_{year}"
        next_m = (m_int % 12) + 1
        next_y = year if m_int < 12 else year + 1
        next_period_title = f"Tháng {next_m:02d}/{next_y}"
        next_period_full = f"THÁNG {next_m:02d} NĂM {next_y}"
    elif period_type == "quarter":
        q_val = str(period_value).upper()
        q_map = {"Q1": [1,2,3], "Q2": [4,5,6], "Q3": [7,8,9], "Q4": [10,11,12], "1": [1,2,3], "2": [4,5,6], "3": [7,8,9], "4": [10,11,12]}
        months = q_map.get(q_val, [7,8,9])
        clean_q = q_val if q_val.startswith("Q") else f"Q{q_val}"
        period_title = f"QUÝ {clean_q[-1]} NĂM {year}"
        period_key = f"Q_{clean_q}_{year}"
        next_q_num = (int(clean_q[-1]) % 4) + 1
        next_period_title = f"Quý {next_q_num} Năm {year if next_q_num > 1 else year + 1}"
        next_period_full = f"QUÝ {next_q_num} NĂM {year if next_q_num > 1 else year + 1}"
    elif period_type == "half_year":
        if "2" in str(period_value) or "CUOI" in str(period_value).upper():
            months = [7,8,9,10,11,12]
            period_title = f"6 THÁNG CUỐI NĂM {year}"
            period_key = f"H2_{year}"
            next_period_title = f"6 Tháng đầu năm {year + 1}"
            next_period_full = f"6 THÁNG ĐẦU NĂM {year + 1}"
        else:
            months = [1,2,3,4,5,6]
            period_title = f"6 THÁNG ĐẦU NĂM {year}"
            period_key = f"H1_{year}"
            next_period_title = f"6 Tháng cuối năm {year}"
            next_period_full = f"6 THÁNG CUỐI NĂM {year}"
    else: # full_year
        months = list(range(1, 13))
        period_title = f"CẢ NĂM {year}"
        period_key = f"Y_{year}"
        next_period_title = f"Năm {year + 1}"
        next_period_full = f"NĂM {year + 1}"

    return {
        "period_type": period_type,
        "period_value": period_value,
        "year": int(year),
        "period_key": period_key,
        "period_title": period_title,
        "next_period_title": next_period_title,
        "next_period_full": next_period_full,
        "months": months
    }

def get_default_hr_data():
    return [
        {"stt": 1, "position": "Văn Phòng", "dinhbien_dc1": 3, "dinhbien_dc2": 3, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 0, "chuyen_dc2": 0, "nghi_dc1": 0, "nghi_dc2": 0, "hientai_dc1": 3, "hientai_dc2": 3},
        {"stt": 2, "position": "Đốc công", "dinhbien_dc1": 4, "dinhbien_dc2": 5, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 0, "chuyen_dc2": 0, "nghi_dc1": 0, "nghi_dc2": 0, "hientai_dc1": 4, "hientai_dc2": 5},
        {"stt": 3, "position": "Ép - TM", "dinhbien_dc1": 27, "dinhbien_dc2": 30, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 0, "chuyen_dc2": 0, "nghi_dc1": 2, "nghi_dc2": 0, "hientai_dc1": 25, "hientai_dc2": 30},
        {"stt": 4, "position": "Lò nung", "dinhbien_dc1": 15, "dinhbien_dc2": 23, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 0, "chuyen_dc2": 0, "nghi_dc1": 0, "nghi_dc2": 0, "hientai_dc1": 15, "hientai_dc2": 23},
        {"stt": 5, "position": "PLĐG", "dinhbien_dc1": 45, "dinhbien_dc2": 33, "tuyenmoi_dc1": 2, "tuyenmoi_dc2": 4, "chuyen_dc1": 5, "chuyen_dc2": 0, "nghi_dc1": 3, "nghi_dc2": 3, "hientai_dc1": 39, "hientai_dc2": 34},
        {"stt": 6, "position": "Xe nâng", "dinhbien_dc1": 6, "dinhbien_dc2": 6, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 0, "chuyen_dc2": 0, "nghi_dc1": 0, "nghi_dc2": 0, "hientai_dc1": 6, "hientai_dc2": 6},
        {"stt": 7, "position": "VSCN", "dinhbien_dc1": 2, "dinhbien_dc2": 2, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 0, "chuyen_dc2": 0, "nghi_dc1": 0, "nghi_dc2": 0, "hientai_dc1": 2, "hientai_dc2": 2},
        {"stt": 8, "position": "Tạo Bột - Tạo Men", "dinhbien_dc1": 62, "dinhbien_dc2": 0, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 0, "chuyen_dc2": 0, "nghi_dc1": 0, "nghi_dc2": 0, "hientai_dc1": 62, "hientai_dc2": 0},
        {"stt": 9, "position": "Trạm Than Hoá Khí", "dinhbien_dc1": 26, "dinhbien_dc2": 0, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 7, "chuyen_dc2": 0, "nghi_dc1": 0, "nghi_dc2": 0, "hientai_dc1": 26, "hientai_dc2": 0},
        {"stt": 10, "position": "Phòng KT - CN", "dinhbien_dc1": 24, "dinhbien_dc2": 0, "tuyenmoi_dc1": 0, "tuyenmoi_dc2": 0, "chuyen_dc1": 0, "chuyen_dc2": 0, "nghi_dc1": 0, "nghi_dc2": 0, "hientai_dc1": 24, "hientai_dc2": 0}
    ]

def get_default_hr_notes():
    return "- DC1 ngày 03/06 chuyển 02 đốc công và 05 công nhân PLĐG sang THK hỗ trợ. Đề nghị P.TC-HC tuyển bổ sung đủ định biên tối thiểu 102 người.\n- DC2 tuyển mới 4 người, nghỉ việc 3 người. Nhân sự cơ bản đáp ứng sản xuất.\n- Bộ phận TB-TM định biên 62/62 đủ nhân sự; Phòng KT-CN định biên 24/24 đủ nhân sự."

def get_default_plan_next_period():
    """Default Plan matching Page 1 of PDF (e.g. Month 9/2026)"""
    return {
        "items": [
            {
                "line": "DC1", "size": "30x60",
                "plan_m2": {"sl_ep": 471744.0, "a1": 416078.0, "a": 32362.0, "b": 13869.0, "recovery_total": 462309.0, "prod_days": 30.0, "avg_per_day": 15410.0, "a_ep": 98.0, "c_ep": 2.0, "huy_ep": 2.0, "stop_time_2mf": 40},
                "plan_pct": {"a1": 90.0, "a": 7.0, "b": 3.0, "recovery_total": 100.0}
            },
            {
                "line": "DC2", "size": "40x80",
                "plan_m2": {"sl_ep": 439488.0, "a1": 387628.0, "a": 0.0, "b": 43070.0, "recovery_total": 430698.0, "prod_days": 30.0, "avg_per_day": 14357.0, "a_ep": 98.0, "c_ep": 2.0, "huy_ep": 2.0, "stop_time_2mf": 40},
                "plan_pct": {"a1": 90.0, "a": 0.0, "b": 10.0, "recovery_total": 100.0}
            }
        ],
        "total_2dc": {
            "plan_m2": {"sl_ep": 911232.0, "a1": 803707.0, "a": 32362.0, "b": 56939.0, "recovery_total": 893007.0, "prod_days": 60.0, "avg_per_day": 14883.0, "a_ep": 98.0, "c_ep": 2.0, "huy_ep": 2.0, "stop_time_2mf": 40},
            "plan_pct": {"a1": 90.0, "a": 3.6, "b": 6.4, "recovery_total": 100.0}
        },
        "notes": [
            "1. Dây chuyền 1: Tổng 30 ngày sản xuất chạy 300x600 mm. Sử dụng BPL Xương PN33 và BPL Men EP17C + GP17A. (chu kỳ ép: 15.6 x 0,72 x 1400 x 0,98 = 15.410 m2/ngày).",
            "2. Dây chuyền 2: Tổng 30 ngày sản xuất kích thước 400x800mm sử dụng BPL Xương PN33 và BPL Men EP17C + PSG17A (Chu kỳ ép: 10,9 x 0,96 x 1400 x 0,98 = 14.357 m2/ngày)."
        ]
    }

def get_default_goals_next_period():
    """Default Goals matching Page 2 of PDF (e.g. Month 9/2026)"""
    return {
        "items": [
            {
                "line": "DC1", "size": "30x60",
                "goal_m2": {"sl_ep": 481179.0, "a1": 434716.0, "a": 23626.0, "b": 14176.0, "recovery_total": 472518.0, "prod_days": 30.0, "avg_per_day": 15751.0, "a_ep": 98.2, "c_ep": 1.80, "huy_ep": 1.80, "stop_time_2mf": 25},
                "goal_pct": {"a1": 92.0, "a": 5.0, "b": 3.0, "recovery_total": 100.0}
            },
            {
                "line": "DC2", "size": "40x80",
                "goal_m2": {"sl_ep": 448278.0, "a1": 404992.0, "a": 0.0, "b": 35217.0, "recovery_total": 440209.0, "prod_days": 30.0, "avg_per_day": 14674.0, "a_ep": 98.2, "c_ep": 1.80, "huy_ep": 1.80, "stop_time_2mf": 25},
                "goal_pct": {"a1": 92.0, "a": 0.0, "b": 8.0, "recovery_total": 100.0}
            }
        ],
        "total_dc2": {
            "goal_m2": {"sl_ep": 448278.0, "a1": 404992.0, "a": 0.0, "b": 35217.0, "recovery_total": 440209.0, "prod_days": 30.0, "avg_per_day": 14674.0, "a_ep": 98.2, "c_ep": 1.80, "huy_ep": 1.80, "stop_time_2mf": 25},
            "goal_pct": {"a1": 92.0, "a": 0.0, "b": 8.0, "recovery_total": 100.0}
        },
        "total_2dc": {
            "goal_m2": {"sl_ep": 929457.0, "a1": 839708.0, "a": 23626.0, "b": 49392.0, "recovery_total": 912726.0, "prod_days": 60.0, "avg_per_day": 15212.0, "a_ep": 98.2, "c_ep": 1.80, "huy_ep": 1.80, "stop_time_2mf": 25},
            "goal_pct": {"a1": 92.0, "a": 2.6, "b": 5.4, "recovery_total": 100.0}
        },
        "department_tasks": [
            {
                "dept": "PXCĐ-NL",
                "tasks": "Lên kế hoạch chi tiết cho việc bảo trì - bảo dưỡng cuối năm. Kết hợp PXSX và P.KHTH kiểm tra các vị trí nhà xưởng bảo đảm mùa mưa không ảnh hưởng chất lượng và máy móc thiết bị."
            },
            {
                "dept": "PXSX",
                "tasks": "Công tác dừng giờ cao điểm (theo quy định mới) kế hoạch kiểm tra định kỳ lớp lót các cối nghiền và lò than xích (đặc biệt là vòng bi các máy nghiền, mô tơ hộp số lò than xích vừa thay). Kết hợp P.KT-CN quản lý, bảo quản hồ men tồn."
            },
            {
                "dept": "P.KT-CN",
                "tasks": "Tiếp tục giám sát, kiểm tra chất lượng nguyên nhiên vật liệu đầu vào ổn định, duy trì xuyên suốt phục vụ sản xuất ổn định. Kiểm tra, giám sát chất lượng sản phẩm 300x600, và 400x800 mm men panson. Kết hợp PXSX và P.KHTH lên kế hoạch chuẩn bị kho bãi, phương án vận chuyển nguyên liệu xương vào mùa mưa phục vụ sản xuất. Kiểm tra các loại nguyên liệu tràng thạch, đất sét phục vụ thay thế nguồn VC02 - VT02 sắp hết. Triển khai chạy BCN bài xương sử dụng tràng thạch PH04 (đã nhập kho 16). Chạy BCN điều chỉnh tỷ lệ đất sét - tràng thạch phục vụ SX lâu dài khi biến động nguồn nguyên liệu."
            },
            {
                "dept": "Phối hợp & Môi trường",
                "tasks": "- Các bộ phận (PXSX - PXCĐ.NL và P.KTCN) gửi kế hoạch chi tiết và gộp chung thành 1 bộ cho P.TGĐ PT.\n- Các phòng ban/ phân xưởng trong toàn Công ty tiếp tục duy trì việc thực hiện, áp dụng (nội bộ) hệ thống Quản lý môi trường ISO 14001:2015. Tuyên truyền công tác bảo vệ môi trường và đôn đốc CBCNV thực hiện ý thức bảo vệ môi trường tại khu vực quản lý."
            }
        ]
    }

def build_form_mau_payload(conn, period_type="month", period_value="8", year=2026):
    conn.row_factory = sqlite3.Row
    init_form_mau_db(conn)
    cur = conn.cursor()
    p_info = resolve_period(period_type, period_value, year)
    months = p_info["months"]
    placeholders = ",".join(["?"] * len(months))

    # 1. SECTION I: SẢN LƯỢNG - THU HỒI TỔNG A1+A+B/ÉP
    q_sec1 = f'''
        SELECT line, size, product_line, data_type,
               SUM(sl_ep) as sl_ep,
               SUM(a1) as a1,
               SUM(a) as a,
               SUM(b) as b,
               SUM(recovery_total) as recovery_total,
               SUM(prod_days) as prod_days,
               AVG(stop_time_2mf) as stop_time_2mf
        FROM data_production_summary
        WHERE month IN ({placeholders}) AND unit = 'm2'
        GROUP BY line, size, data_type
        ORDER BY line, size, data_type DESC
    '''
    sec1_raw = [dict(r) for r in cur.execute(q_sec1, months).fetchall()]

    target_groups = [
        {"line": "DC1", "size": "30x60", "name": "DC1 30x60"},
        {"line": "DC2", "size": "50x50", "name": "DC2 50x50"},
        {"line": "DC2", "size": "60x60", "name": "DC2 60x60"},
        {"line": "DC2", "size": "40x80", "name": "DC2 40x80"}
    ]

    sec1_items = []
    total_plan_dc2 = {"sl_ep": 0, "a1": 0, "a": 0, "b": 0, "recovery_total": 0, "prod_days": 0, "stop_time_2mf": 0}
    total_actual_dc2 = {"sl_ep": 0, "a1": 0, "a": 0, "b": 0, "recovery_total": 0, "prod_days": 0, "stop_time_2mf": 0}
    total_plan_all = {"sl_ep": 0, "a1": 0, "a": 0, "b": 0, "recovery_total": 0, "prod_days": 0, "stop_time_2mf": 0}
    total_actual_all = {"sl_ep": 0, "a1": 0, "a": 0, "b": 0, "recovery_total": 0, "prod_days": 0, "stop_time_2mf": 0}

    for tg in target_groups:
        plan_row = next((r for r in sec1_raw if r["line"] == tg["line"] and r["size"] == tg["size"] and "kế hoạch" in (r["data_type"] or "").lower()), None)
        actual_row = next((r for r in sec1_raw if r["line"] == tg["line"] and r["size"] == tg["size"] and "thực hiện" in (r["data_type"] or "").lower()), None)

        def make_row_data(r):
            if not r:
                return {"sl_ep": 0, "a1": 0, "a": 0, "b": 0, "recovery_total": 0, "pct_a1": 0, "pct_a": 0, "pct_b": 0, "prod_days": 0, "avg_per_day": 0, "a_ep": 0, "c_ep": 0, "huy_ep": 0, "stop_time_2mf": 0}
            sl_ep = r["sl_ep"] or 0
            a1 = r["a1"] or 0
            a = r["a"] or 0
            b = r["b"] or 0
            rec = r["recovery_total"] or (a1 + a + b)
            days = r["prod_days"] or 0
            stop_mf = r["stop_time_2mf"] or 0

            return {
                "sl_ep": sl_ep,
                "a1": a1,
                "a": a,
                "b": b,
                "recovery_total": rec,
                "pct_a1": (a1 / rec * 100) if rec > 0 else 0,
                "pct_a": (a / rec * 100) if rec > 0 else 0,
                "pct_b": (b / rec * 100) if rec > 0 else 0,
                "prod_days": days,
                "avg_per_day": (rec / days) if days > 0 else 0,
                "a_ep": (rec / sl_ep * 100) if sl_ep > 0 else 0,
                "c_ep": (b / sl_ep * 100) if sl_ep > 0 else 0,
                "huy_ep": ((sl_ep - rec) / sl_ep * 100) if sl_ep > 0 else 0,
                "stop_time_2mf": stop_mf
            }

        p_data = make_row_data(plan_row)
        a_data = make_row_data(actual_row)

        if tg["line"] == "DC2":
            for k in total_plan_dc2: total_plan_dc2[k] += p_data.get(k, 0)
            for k in total_actual_dc2: total_actual_dc2[k] += a_data.get(k, 0)
        for k in total_plan_all: total_plan_all[k] += p_data.get(k, 0)
        for k in total_actual_all: total_actual_all[k] += a_data.get(k, 0)

        diff_m2 = {
            "sl_ep": a_data["sl_ep"] - p_data["sl_ep"],
            "a1": a_data["a1"] - p_data["a1"],
            "a": a_data["a"] - p_data["a"],
            "b": a_data["b"] - p_data["b"],
            "recovery_total": a_data["recovery_total"] - p_data["recovery_total"],
            "prod_days": a_data["prod_days"] - p_data["prod_days"],
            "avg_per_day": a_data["avg_per_day"] - p_data["avg_per_day"],
            "stop_time_2mf": a_data["stop_time_2mf"] - p_data["stop_time_2mf"]
        }

        # IMPORTANT: Rate difference according to user formula (Actual % - Plan %) for A1, A, B, A/ép, C/ép, Huỷ/ép
        rate_pct = {
            "sl_ep": (a_data["sl_ep"] / p_data["sl_ep"] * 100) if p_data["sl_ep"] > 0 else 0,
            "a1": a_data["pct_a1"] - p_data["pct_a1"], # e.g. 93.81 - 90.00 = +3.81%
            "a": a_data["pct_a"] - p_data["pct_a"],     # e.g. 2.64 - 7.00 = -4.36%
            "b": a_data["pct_b"] - p_data["pct_b"],     # e.g. 3.54 - 3.00 = +0.54%
            "recovery_total": (a_data["recovery_total"] / p_data["recovery_total"] * 100) if p_data["recovery_total"] > 0 else 0,
            "a_ep": a_data["a_ep"] - p_data["a_ep"],
            "c_ep": a_data["c_ep"] - p_data["c_ep"],
            "huy_ep": a_data["huy_ep"] - p_data["huy_ep"]
        }

        sec1_items.append({
            "group_name": tg["name"],
            "line": tg["line"],
            "size": tg["size"],
            "plan": p_data,
            "actual": a_data,
            "diff_m2": diff_m2,
            "rate_pct": rate_pct
        })

    def finalize_total_group(p, a, name):
        p_rec = p["recovery_total"] or (p["a1"] + p["a"] + p["b"])
        a_rec = a["recovery_total"] or (a["a1"] + a["a"] + a["b"])
        p_days = p["prod_days"]
        a_days = a["prod_days"]
        p_ep = p["sl_ep"]
        a_ep = a["sl_ep"]

        p_res = {
            "sl_ep": p_ep, "a1": p["a1"], "a": p["a"], "b": p["b"], "recovery_total": p_rec,
            "pct_a1": (p["a1"] / p_rec * 100) if p_rec > 0 else 0,
            "pct_a": (p["a"] / p_rec * 100) if p_rec > 0 else 0,
            "pct_b": (p["b"] / p_rec * 100) if p_rec > 0 else 0,
            "prod_days": p_days, "avg_per_day": (p_rec / p_days) if p_days > 0 else 0,
            "a_ep": (p_rec / p_ep * 100) if p_ep > 0 else 0,
            "c_ep": (p["b"] / p_ep * 100) if p_ep > 0 else 0,
            "huy_ep": ((p_ep - p_rec) / p_ep * 100) if p_ep > 0 else 0,
            "stop_time_2mf": p["stop_time_2mf"]
        }
        a_res = {
            "sl_ep": a_ep, "a1": a["a1"], "a": a["a"], "b": a["b"], "recovery_total": a_rec,
            "pct_a1": (a["a1"] / a_rec * 100) if a_rec > 0 else 0,
            "pct_a": (a["a"] / a_rec * 100) if a_rec > 0 else 0,
            "pct_b": (a["b"] / a_rec * 100) if a_rec > 0 else 0,
            "prod_days": a_days, "avg_per_day": (a_rec / a_days) if a_days > 0 else 0,
            "a_ep": (a_rec / a_ep * 100) if a_ep > 0 else 0,
            "c_ep": (a["b"] / a_ep * 100) if a_ep > 0 else 0,
            "huy_ep": ((a_ep - a_rec) / a_ep * 100) if a_ep > 0 else 0,
            "stop_time_2mf": a["stop_time_2mf"]
        }
        return {
            "name": name,
            "plan": p_res,
            "actual": a_res,
            "diff_m2": {
                "sl_ep": a_res["sl_ep"] - p_res["sl_ep"],
                "a1": a_res["a1"] - p_res["a1"],
                "a": a_res["a"] - p_res["a"],
                "b": a_res["b"] - p_res["b"],
                "recovery_total": a_res["recovery_total"] - p_res["recovery_total"],
                "prod_days": a_res["prod_days"] - p_res["prod_days"],
                "avg_per_day": a_res["avg_per_day"] - p_res["avg_per_day"],
                "stop_time_2mf": a_res["stop_time_2mf"] - p_res["stop_time_2mf"]
            },
            "rate_pct": {
                "sl_ep": (a_res["sl_ep"] / p_res["sl_ep"] * 100) if p_res["sl_ep"] > 0 else 0,
                "a1": a_res["pct_a1"] - p_res["pct_a1"],
                "a": a_res["pct_a"] - p_res["pct_a"],
                "b": a_res["pct_b"] - p_res["pct_b"],
                "recovery_total": (a_res["recovery_total"] / p_res["recovery_total"] * 100) if p_res["recovery_total"] > 0 else 0,
                "a_ep": a_res["a_ep"] - p_res["a_ep"],
                "c_ep": a_res["c_ep"] - p_res["c_ep"],
                "huy_ep": a_res["huy_ep"] - p_res["huy_ep"]
            }
        }

    total_dc2_obj = finalize_total_group(total_plan_dc2, total_actual_dc2, "TỔNG DC 2")
    total_2dc_obj = finalize_total_group(total_plan_all, total_actual_all, "TỔNG 2 DC")

    # 2. SECTION II: THƯƠNG HIỆU & CƠ CẤU MEN
    q_sec2 = f'''
        SELECT line, size, glaze_type, brand_name,
               SUM(CASE WHEN grade = 'A1' THEN quantity_m2 ELSE 0 END) as a1_m2,
               SUM(CASE WHEN grade = 'A' THEN quantity_m2 ELSE 0 END) as a_m2,
               SUM(CASE WHEN grade = 'B' THEN quantity_m2 ELSE 0 END) as b_m2,
               SUM(quantity_m2) as total_m2
        FROM data_brand_production
        WHERE month IN ({placeholders})
        GROUP BY line, size, glaze_type, brand_name
        ORDER BY line, size, total_m2 DESC
    '''
    sec2_raw = [dict(r) for r in cur.execute(q_sec2, months).fetchall()]

    dc1_brands = [r for r in sec2_raw if r["line"] == "DC1"]
    sum_dc1_brands = {
        "a1": sum(r["a1_m2"] for r in dc1_brands),
        "a": sum(r["a_m2"] for r in dc1_brands),
        "b": sum(r["b_m2"] for r in dc1_brands),
        "total": sum(r["total_m2"] for r in dc1_brands)
    }

    q_glaze = f'''
        SELECT glaze_type, size,
               SUM(quantity_m2) as prod_m2,
               COUNT(DISTINCT month) as months_active
        FROM data_brand_production
        WHERE month IN ({placeholders}) AND line = 'DC1'
        GROUP BY glaze_type, size
        ORDER BY prod_m2 DESC
    '''
    dc1_glazes = [dict(r) for r in cur.execute(q_glaze, months).fetchall()]
    sum_dc1_glazes = sum(r["prod_m2"] for r in dc1_glazes)
    for g in dc1_glazes:
        g["pct"] = (g["prod_m2"] / sum_dc1_glazes * 100) if sum_dc1_glazes > 0 else 0

    # Split DC2 50x50 into Men Bóng and Men Sugar Sân Vườn
    dc2_50x50_all = [r for r in sec2_raw if r["line"] == "DC2" and r["size"] == "50x50"]
    dc2_50x50_bong = [r for r in dc2_50x50_all if "bóng" in (r["glaze_type"] or "").lower()]
    dc2_50x50_sugar = [r for r in dc2_50x50_all if "sugar" in (r["glaze_type"] or "").lower() or "vườn" in (r["glaze_type"] or "").lower()]
    
    # In case there are items not matched into bong or sugar
    other_50x50 = [r for r in dc2_50x50_all if r not in dc2_50x50_bong and r not in dc2_50x50_sugar]
    if other_50x50:
        dc2_50x50_bong.extend(other_50x50)

    def calc_brand_subtotal(rows_list):
        return {
            "a1": sum(r["a1_m2"] for r in rows_list),
            "a": sum(r["a_m2"] for r in rows_list),
            "b": sum(r["b_m2"] for r in rows_list),
            "total": sum(r["total_m2"] for r in rows_list)
        }

    sum_50x50_bong = calc_brand_subtotal(dc2_50x50_bong)
    sum_50x50_sugar = calc_brand_subtotal(dc2_50x50_sugar)
    sum_50x50_total = calc_brand_subtotal(dc2_50x50_all)

    dc2_60x60_brands = [r for r in sec2_raw if r["line"] == "DC2" and r["size"] == "60x60"]
    sum_60x60_total = calc_brand_subtotal(dc2_60x60_brands)

    dc2_40x80_brands = [r for r in sec2_raw if r["line"] == "DC2" and r["size"] == "40x80"]
    sum_40x80_total = calc_brand_subtotal(dc2_40x80_brands)

    # 3. SECTION III: TIÊU HAO VẬT TƯ (DC1 30x60, DC2 50x50, 60x60, 40x80)
    q_sec3 = f'''
        SELECT line, size, material_name, unit,
               AVG(norm_value) as norm_value,
               SUM(used_qty) as used_qty,
               SUM(prod_qty) as prod_qty,
               AVG(actual_rate) as actual_rate,
               SUM(diff_qty) as diff_qty
        FROM data_material_consumption
        WHERE month IN ({placeholders})
        GROUP BY line, size, material_name
        ORDER BY line, size, id ASC
    '''
    sec3_raw = [dict(r) for r in cur.execute(q_sec3, months).fetchall()]
    mat_dc1_30x60 = [r for r in sec3_raw if r["line"] == "DC1" and r["size"] == "30x60"]
    mat_dc2_50x50 = [r for r in sec3_raw if r["line"] == "DC2" and r["size"] == "50x50"]
    mat_dc2_60x60 = [r for r in sec3_raw if r["line"] == "DC2" and r["size"] == "60x60"]
    mat_dc2_40x80 = [r for r in sec3_raw if r["line"] == "DC2" and r["size"] == "40x80"]

    # 4. SECTION IV: SỬ DỤNG THAN (Structured per size and total DC2, total 2DC)
    q_sec4 = f'''
        SELECT line, size, firing_type,
               SUM(issued_weight) as issued_weight,
               SUM(ash_weight) as ash_weight,
               SUM(compensation_weight) as compensation_weight,
               SUM(excess_ash_weight) as excess_ash_weight,
               SUM(total_used_weight) as total_used_weight,
               SUM(production_m2) as production_m2
        FROM data_coal_consumption
        WHERE month IN ({placeholders})
        GROUP BY line, size, firing_type
        ORDER BY line, size
    '''
    sec4_raw = [dict(r) for r in cur.execute(q_sec4, months).fetchall()]

    def summarize_coal_group(rows_list):
        issued = sum(r["issued_weight"] or 0 for r in rows_list)
        ash = sum(r["ash_weight"] or 0 for r in rows_list)
        comp = sum(r["compensation_weight"] or 0 for r in rows_list)
        total_used = sum(r["total_used_weight"] or 0 for r in rows_list)
        prod_m2 = sum(r["production_m2"] or 0 for r in rows_list)
        return {
            "issued_weight": issued,
            "ash_weight": ash,
            "compensation_weight": comp,
            "total_used_weight": total_used,
            "production_m2": prod_m2,
            "rate_kg_m2": (total_used / prod_m2) if prod_m2 > 0 else 0
        }

    coal_dc1_30x60 = summarize_coal_group([r for r in sec4_raw if r["line"] == "DC1"])
    coal_dc2_50x50 = summarize_coal_group([r for r in sec4_raw if r["line"] == "DC2" and r["size"] == "50x50"])
    coal_dc2_60x60 = summarize_coal_group([r for r in sec4_raw if r["line"] == "DC2" and r["size"] == "60x60"])
    coal_dc2_40x80 = summarize_coal_group([r for r in sec4_raw if r["line"] == "DC2" and r["size"] == "40x80"])
    coal_dc2_total = summarize_coal_group([r for r in sec4_raw if r["line"] == "DC2"])
    coal_all_total = summarize_coal_group(sec4_raw)

    coal_table_rows = [
        {"stt": 1, "name": "Dây chuyền số 1 (300x600)", "data": coal_dc1_30x60, "eval": "Đạt định mức khoán ✓", "is_total": False},
        {"stt": 2, "name": "DC2 - Kích thước 500x500", "data": coal_dc2_50x50, "eval": "Ổn định", "is_total": False},
        {"stt": 3, "name": "DC2 - Kích thước 600x600", "data": coal_dc2_60x60, "eval": "Ổn định", "is_total": False},
        {"stt": 4, "name": "DC2 - Kích thước 400x800", "data": coal_dc2_40x80, "eval": "Ổn định", "is_total": False},
        {"stt": "", "name": "TỔNG SỬ DỤNG DÂY CHUYỀN 2", "data": coal_dc2_total, "eval": "", "is_total": True},
        {"stt": "", "name": "TỔNG SỬ DỤNG 2 DÂY CHUYỀN", "data": coal_all_total, "eval": "Đạt kế hoạch năm ✓", "is_total": True}
    ]

    # 5, 6, 7, 8: FETCH CUSTOM DATA FROM DB OR DEFAULTS
    saved_custom = cur.execute("SELECT * FROM report_form_mau_custom WHERE period_key = ?", (p_info["period_key"],)).fetchone()
    
    hr_data = get_default_hr_data()
    hr_notes = get_default_hr_notes()
    plan_data = get_default_plan_next_period()
    goals_data = get_default_goals_next_period()
    eval_text = "- Toàn bộ 2 dây chuyền trong kỳ cơ bản hoàn thành các chỉ tiêu sản xuất, chất lượng và thu hồi.\n- Các sự cố dừng máy điện lưới, bảo dưỡng cơ điện và chuyển đổi kích thước được xử lý nhanh chóng, an toàn tuyệt đối.\n- Đề nghị Ban Giám Đốc và các Phòng ban tiếp tục hỗ trợ nhân sự và cung ứng vật tư kịp thời để phân xưởng hoàn thành vượt mức mục tiêu kỳ tới."
    signatures = {
        "date_str": "Đồng Nai, ngày 28 tháng 08 năm 2026",
        "signer_1_title": "TỔNG GIÁM ĐỐC", "signer_1_name": "",
        "signer_2_title": "P.TGĐ PT", "signer_2_name": "",
        "signer_3_title": "PXCĐ-NL", "signer_3_name": "",
        "signer_4_title": "PXSX", "signer_4_name": "",
        "signer_5_title": "P.KT-CN", "signer_5_name": "",
        "signer_6_title": "Người lập", "signer_6_name": ""
    }

    if saved_custom:
        s_dict = dict(saved_custom)
        if s_dict.get("hr_data"):
            try: hr_data = json.loads(s_dict["hr_data"])
            except: pass
        if s_dict.get("notes_data"):
            try:
                n_data = json.loads(s_dict["notes_data"])
                hr_notes = n_data.get("hr_notes", hr_notes)
            except: pass
        if s_dict.get("plan_data"):
            try: plan_data = json.loads(s_dict["plan_data"])
            except: pass
        if s_dict.get("goals_data"):
            try: goals_data = json.loads(s_dict["goals_data"])
            except: pass
        if s_dict.get("evaluation_data"):
            eval_text = s_dict["evaluation_data"]
        if s_dict.get("signatures_data"):
            try: signatures = json.loads(s_dict["signatures_data"])
            except: pass

    return {
        "period_info": p_info,
        "section_1_production": {
            "items": sec1_items,
            "total_dc2": total_dc2_obj,
            "total_2dc": total_2dc_obj,
            "notes": "- DC1 chạy ổn định chu kỳ ép 15.5 lần/phút.\n- DC2 sản xuất các dòng kích thước 50x50, 60x60 và 40x80 đảm bảo kế hoạch."
        },
        "section_2_brands": {
            "dc1_30x60": dc1_brands,
            "dc1_30x60_sum": sum_dc1_brands,
            "dc1_glazes": dc1_glazes,
            "dc2_50x50": {
                "bong": dc2_50x50_bong,
                "sugar": dc2_50x50_sugar,
                "sum_bong": sum_50x50_bong,
                "sum_sugar": sum_50x50_sugar,
                "sum_total": sum_50x50_total
            },
            "dc2_60x60": dc2_60x60_brands,
            "dc2_60x60_sum": sum_60x60_total,
            "dc2_40x80": dc2_40x80_brands,
            "dc2_40x80_sum": sum_40x80_total
        },
        "section_3_materials": {
            "dc1_30x60": mat_dc1_30x60,
            "dc2_50x50": mat_dc2_50x50,
            "dc2_60x60": mat_dc2_60x60,
            "dc2_40x80": mat_dc2_40x80
        },
        "section_4_coal": {
            "rows": coal_table_rows,
            "dc1_30x60": coal_dc1_30x60,
            "dc2_50x50": coal_dc2_50x50,
            "dc2_60x60": coal_dc2_60x60,
            "dc2_40x80": coal_dc2_40x80,
            "total_dc2": coal_dc2_total,
            "total_2dc": coal_all_total
        },
        "section_5_hr": {
            "table": hr_data,
            "notes": hr_notes
        },
        "section_6_plan": {
            "next_title": p_info["next_period_title"],
            "next_full_title": p_info["next_period_full"],
            "data": plan_data
        },
        "section_7_goals": {
            "next_title": p_info["next_period_title"],
            "next_full_title": p_info["next_period_full"],
            "data": goals_data
        },
        "section_8_evaluation": {
            "content": eval_text,
            "signatures": signatures
        }
    }

def save_form_mau_custom_data(conn, body):
    conn.row_factory = sqlite3.Row
    init_form_mau_db(conn)
    cur = conn.cursor()
    p_type = body.get("period_type", "month")
    p_val = str(body.get("period_value", "8"))
    year = int(body.get("year", 2026))
    p_info = resolve_period(p_type, p_val, year)
    p_key = p_info["period_key"]

    hr_json = json.dumps(body.get("hr_data", [])) if "hr_data" in body else None
    plan_json = json.dumps(body.get("plan_data", {})) if "plan_data" in body else None
    goals_json = json.dumps(body.get("goals_data", {})) if "goals_data" in body else None
    notes_json = json.dumps(body.get("notes_data", {})) if "notes_data" in body else None
    eval_text = body.get("evaluation_data") if "evaluation_data" in body else None
    sig_json = json.dumps(body.get("signatures_data", {})) if "signatures_data" in body else None

    cur.execute('''
        INSERT INTO report_form_mau_custom (
            period_key, period_type, period_value, year,
            hr_data, plan_data, goals_data, notes_data, evaluation_data, signatures_data, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(period_key) DO UPDATE SET
            hr_data = COALESCE(excluded.hr_data, report_form_mau_custom.hr_data),
            plan_data = COALESCE(excluded.plan_data, report_form_mau_custom.plan_data),
            goals_data = COALESCE(excluded.goals_data, report_form_mau_custom.goals_data),
            notes_data = COALESCE(excluded.notes_data, report_form_mau_custom.notes_data),
            evaluation_data = COALESCE(excluded.evaluation_data, report_form_mau_custom.evaluation_data),
            signatures_data = COALESCE(excluded.signatures_data, report_form_mau_custom.signatures_data),
            updated_at = CURRENT_TIMESTAMP
    ''', (p_key, p_type, p_val, year, hr_json, plan_json, goals_json, notes_json, eval_text, sig_json))
    conn.commit()
    return {"success": True, "period_key": p_key}

def parse_form_mau_excel_upload(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    logs = []
    parsed_data = {}

    sheet_names = wb.sheetnames
    logs.append(f"Tìm thấy các sheet: {', '.join(sheet_names)}")

    target_sheet = None
    for name in ["FORM MẪU", "FORM MAU", "KẾ HOẠCH", "KE HOACH", "BÁO CÁO", "Sheet1"]:
        match = next((s for s in sheet_names if name.lower() in s.lower()), None)
        if match:
            target_sheet = wb[match]
            logs.append(f"Sử dụng sheet: '{match}'")
            break
    if not target_sheet:
        target_sheet = wb.active
        logs.append(f"Sử dụng active sheet: '{target_sheet.title}'")

    rows = list(target_sheet.iter_rows(values_only=True))
    logs.append(f"Đọc tổng cộng {len(rows)} dòng từ file Excel.")

    # 1. Parse Section V: Nhân sự
    hr_rows = []
    found_hr = False
    for i, r in enumerate(rows):
        if not r or not any(r): continue
        r_str = " ".join([str(c) for c in r if c is not None]).lower()
        if "tình hình nhân sự" in r_str or "định biên" in r_str or ("đốc công" in r_str and "plđg" in r_str):
            found_hr = True
            for j in range(i + 1, min(i + 25, len(rows))):
                sub_r = rows[j]
                if not sub_r or not any(sub_r): continue
                stt_val = str(sub_r[0] or sub_r[1] or "").strip()
                pos_val = str(sub_r[1] or sub_r[2] or "").strip()
                if any(kw in pos_val.lower() for kw in ["văn phòng", "đốc công", "ép", "lò nung", "plđg", "xe nâng", "vscn", "tạo bột", "than hoá", "phòng kt"]):
                    nums = [float(c) if isinstance(c, (int, float)) else 0 for c in sub_r if isinstance(c, (int, float))]
                    hr_rows.append({
                        "stt": len(hr_rows) + 1,
                        "position": pos_val,
                        "dinhbien_dc1": nums[0] if len(nums) > 0 else 0,
                        "dinhbien_dc2": nums[1] if len(nums) > 1 else 0,
                        "tuyenmoi_dc1": nums[2] if len(nums) > 2 else 0,
                        "tuyenmoi_dc2": nums[3] if len(nums) > 3 else 0,
                        "chuyen_dc1": nums[4] if len(nums) > 4 else 0,
                        "chuyen_dc2": nums[5] if len(nums) > 5 else 0,
                        "nghi_dc1": nums[6] if len(nums) > 6 else 0,
                        "nghi_dc2": nums[7] if len(nums) > 7 else 0,
                        "hientai_dc1": nums[8] if len(nums) > 8 else 0,
                        "hientai_dc2": nums[9] if len(nums) > 9 else 0
                    })
            break

    if hr_rows:
        parsed_data["hr_data"] = hr_rows
        logs.append(f"✓ Trích xuất thành công {len(hr_rows)} dòng nhân sự.")

    return {
        "success": True,
        "logs": logs,
        "parsed_data": parsed_data
    }
